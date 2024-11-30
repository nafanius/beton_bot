import ezsheets
import openpyxl
import re
from datetime import time

# sheets = ezsheets.Spreadsheet('./credentials.json')
# spreadsheets = sheets.listSpreadsheets()
#
# for spreadsheet in spreadsheets:
#     print(spreadsheet)


def get_from_google_sheet():
    ss = ezsheets.Spreadsheet('Tydz 49.2024')
    print(ezsheets.listSpreadsheets())
    ss.downloadAsExcel()




wb = openpyxl.load_workbook('Tydz_49.2024.xlsx')
print(wb.sheetnames)
sheet = wb[wb.sheetnames[0]]
print(sheet)



def form_lista():
    lista = []

    def fill_list(time_str, row, column):
        if isinstance(time_str, time):
            lista.append((time_str, (sheet.cell(row=row, column=column - 1).value).strip()))


    for row_in_file in range(1, sheet.max_row):
        c = sheet.cell(row=row_in_file, column=3).value
        if "zawodzie 2 " in str(c).lower() or "zawodzie 1 " in str(c).lower():
            for row_in_list_time in range(15):
                fill_list(sheet.cell(row=row_in_file+row_in_list_time, column=12).value, row_in_file+row_in_list_time, 12)
                fill_list(sheet.cell(row=row_in_file+row_in_list_time, column=15).value, row_in_file+row_in_list_time, 15)

    lista = sorted(lista, key=lambda event: event[0])

    return lista




